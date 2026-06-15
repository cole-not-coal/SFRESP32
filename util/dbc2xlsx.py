import os
import sys
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

try:
	# Tkinter is in the standard library and provides a simple file picker
	import tkinter as tk
	from tkinter import filedialog
except Exception:
	tk = None
	filedialog = None


try:
	import cantools
except ImportError:
	cantools = None


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DBC = os.path.join(SCRIPT_DIR, "SFR_V3_4.dbc")
# This is the template workbook that will never be modified in-place.
DEFAULT_TEMPLATE_XLSX = os.path.join(SCRIPT_DIR, "CAN_Loading_Template.xlsx")


def load_dbc(path: str):
	if cantools is None:
		print("Error: 'cantools' library is required to parse DBC files.")
		print("Install it with: pip install cantools")
		sys.exit(1)

	if not os.path.exists(path):
		print(f"Error: DBC file not found: {path}")
		sys.exit(1)

	try:
		# Use non-strict loading so placeholder/overlapping signals (e.g. "Blank")
		# do not cause the entire DBC to be rejected.
		db = cantools.database.load_file(path, strict=False)
	except Exception as exc:
		print(f"Failed to load DBC: {exc}")
		sys.exit(1)

	return db


def build_main_bus_sheet(db) -> pd.DataFrame:
	"""Create the 'Main BUS' sheet.

	Columns are chosen to match what decodeCAN.py expects:
	- ID
	- Name
	- Description
	"""

	rows = []
	for msg in db.messages:
		msg_id = msg.frame_id
		name = msg.name or ""
		desc = msg.comment or ""
		length_bytes = getattr(msg, "length", None)
		# Sender is the node(s) listed as message senders in the DBC
		senders = getattr(msg, "senders", None) or []
		sender_str = ", ".join(str(s) for s in senders) if senders else ""
		# Receiver is the union of all signal receivers in the message
		receiver_set: set[str] = set()
		for sig in msg.signals:
			for r in getattr(sig, "receivers", None) or []:
				receiver_set.add(str(r))
		receiver_str = ", ".join(sorted(receiver_set)) if receiver_set else ""

		rows.append(
			{
				"ID": f"0x{msg_id:X}",
				"Name": name,
				"Description": desc,
				"Length (bytes)": length_bytes,
				"Sender": sender_str,
				"Receiver": receiver_str,
			}
		)

	return pd.DataFrame(
		rows,
		columns=["ID", "Name", "Description", "Length (bytes)", "Sender", "Receiver"],
	)


def byte_order_label(signal) -> str:
	"""Return a string for Byte Order column compatible with decodeCAN.py.

	decodeCAN.py checks for "MSB" in the string to detect big-endian.
	"""

	order = getattr(signal, "byte_order", None)
	if order is None:
		return ""

	order_str = str(order).lower()
	if "big" in order_str:
		return "MSB_FIRST"  # contains "MSB" so decodeCAN treats as big-endian
	if "little" in order_str:
		return "LSB_FIRST"
	return order_str.upper()


def original_bit_offset_from_dbc(signal) -> int | None:
	"""Recover the sequential bit offset used in the Excel sheet.

	The MATLAB CANDBGenerator.m script builds the DBC start bit from a
	sequential bit counter ("bitcount") as follows:

	- Little endian:   start = bitcount
	- Big endian:      start = (7 - mod(bitcount,8)) + 8*floor(bitcount/8)

	Here we invert that mapping so that the Excel "start bit" column
	contains the original sequential bit offset (bitcount), which is
	what other tooling expects.
	"""

	start = getattr(signal, "start", None)
	if start is None:
		return None

	try:
		s = int(start)
	except (TypeError, ValueError):
		return None

	order = str(getattr(signal, "byte_order", "little_endian")).lower()
	if "big" in order:
		byte_index = s // 8
		bit_in_byte = s % 8
		# Invert MSBPos = (7 - mod(bitcount,8)) + 8*floor(bitcount/8)
		return (7 - bit_in_byte) + 8 * byte_index
	else:
		# Little endian: DBC start bit already equals bitcount
		return s


def build_main_bus_message_sheet(db) -> pd.DataFrame:
	"""Create the 'Main BUS Message' sheet.

	Column names and types are aligned with decodeCAN.py expectations.
	"""

	rows = []

	for msg in db.messages:
		msg_id = msg.frame_id
		msg_id_str = f"0x{msg_id:X}"
		# Overall frame length in bytes for this message, same for all its signals
		length_bytes = getattr(msg, "length", None)

		for sig in msg.signals:
			# Ignore placeholder signals explicitly named "blank"
			if (getattr(sig, "name", "") or "").strip().lower() == "blank":
				continue

			name = sig.name or ""
			desc = sig.comment or ""

			start_bit = original_bit_offset_from_dbc(sig)
			length = getattr(sig, "length", None)
			gain = getattr(sig, "scale", 1.0)
			offset = getattr(sig, "offset", 0.0)
			is_signed = getattr(sig, "is_signed", False)
			unit = getattr(sig, "unit", "") or ""

			signed_str = "YES" if is_signed else "NO"
			byte_order = byte_order_label(sig)

			rows.append(
				{
					"ID": msg_id_str,
					"Name": name,
					"Description": desc,
					"start bit": start_bit,
					"length (bits)": length,
					"Length (bytes)": length_bytes,
					"Gain": gain,
					"Offset": offset,
					"Signed": signed_str,
					"Byte Order": byte_order,
					"Unit": unit,
				}
			)

	columns = [
		"ID",
		"Name",
		"Description",
		"start bit",
		"length (bits)",
		"Length (bytes)",
		"Gain",
		"Offset",
		"Signed",
		"Byte Order",
		"Unit",
	]

	return pd.DataFrame(rows, columns=columns)


def _find_header_row(ws, required_keywords: list[str]) -> Optional[int]:
	"""Find the header row index (1-based) that contains all required keywords.

	Matching is case-insensitive and based on substring presence in the row.
	"""

	required_lower = [k.lower() for k in required_keywords]
	for r in range(1, ws.max_row + 1):
		values = [str(c.value).strip() for c in ws[r] if c.value is not None]
		row_str = " ".join(values).lower()
		if all(k in row_str for k in required_lower):
			return r
	return None


def _map_headers(ws, header_row: int, column_names: list[str], aliases: dict[str, list[str]] | None = None) -> dict[str, int]:
	"""Map logical column names to Excel column indices based on a header row.

	aliases: mapping from logical name -> list of alternative header strings.
	Returns a mapping {logical_name: column_index} for found headers.
	"""

	if aliases is None:
		aliases = {}

	name_to_col: dict[str, int] = {}
	for cell in ws[header_row]:
		if cell.value is None:
			continue
		helper = str(cell.value).strip().lower()
		for logical in column_names:
			alts = [logical] + aliases.get(logical, [])
			for alt in alts:
				if helper == alt.lower():
					name_to_col[logical] = cell.column
					break

	return name_to_col


def _populate_sheet(ws, header_row: int, header_map: dict[str, int], rows: list[dict[str, object]]):
	"""Write rows under the header row and trim any extra rows below the last one.

	Only columns present in header_map are written; other columns/content are untouched.
	To avoid issues with merged cells in the body area, any merged
	regions that extend below the header row are unmerged in the
	output workbook copy before writing.
	"""

	start_row = header_row + 1

	# Unmerge any body-area merged cells so we can safely write data
	body_merges = [rng for rng in list(ws.merged_cells.ranges) if rng.max_row >= start_row]
	for rng in body_merges:
		ws.unmerge_cells(str(rng))

	# Write all rows
	for i, data in enumerate(rows):
		excel_row = start_row + i
		row_mode = data.get("__row_mode") if isinstance(data, dict) else None

		# Separator row: fully blank
		if not data:
			for cell in ws[excel_row]:
				cell.value = None
			continue

		# Padding row: clear entire row, then write only provided fields
		if row_mode == "padding":
			for cell in ws[excel_row]:
				cell.value = None
			for logical, col_idx in header_map.items():
				if logical in data:
					ws.cell(row=excel_row, column=col_idx, value=data[logical])
			continue

		# Normal data row: only touch mapped columns
		for logical, col_idx in header_map.items():
			if logical in data:
				ws.cell(row=excel_row, column=col_idx, value=data[logical])

	last_row = start_row + len(rows) - 1 if rows else header_row
	# Delete any rows below the last populated one (to remove stale template data)
	if ws.max_row > last_row:
		n_delete = ws.max_row - last_row
		ws.delete_rows(last_row + 1, n_delete)


def _style_message_blocks(ws, header_row: int, header_map: dict[str, int]):
	"""Apply borders around each message block on the message sheet.

	- Thin borders on all cells
	- Thick outer border around each message (grouped by ID)
	- Columns A through N (1..14) are styled
	"""

	id_col = header_map.get("ID")
	if not id_col:
		return

	start_row = header_row + 1
	max_row = ws.max_row
	if max_row < start_row:
		return

	# Determine row ranges for each message.
	# A message starts where ID is non-empty. It continues through
	# subsequent non-blank rows (even if ID column is blank) until a
	# completely blank separator row or the next non-empty ID row.
	message_ranges: list[tuple[int, int]] = []
	cur_start: int | None = None
	last_content_row: int | None = None

	for r in range(start_row, max_row + 1):
		row_cells = ws[r]
		is_blank_row = all(
			(cell.value is None or str(cell.value).strip() == "") for cell in row_cells
		)
		id_val = ws.cell(row=r, column=id_col).value

		if id_val not in (None, ""):
			# New message header row
			if cur_start is not None and last_content_row is not None:
				# Close previous message before starting new one
				message_ranges.append((cur_start, last_content_row))
			cur_start = r
			last_content_row = r
		elif cur_start is not None and not is_blank_row:
			# Continuation row of current message
			last_content_row = r
		# Blank rows while cur_start is set are treated as separators but
		# finalization happens when the next header is seen or at the end.

	# Close last message
	if cur_start is not None and last_content_row is not None:
		message_ranges.append((cur_start, last_content_row))

	if not message_ranges:
		return

	thin = Side(style="thin", color="000000")
	thick = Side(style="thick", color="000000")

	for start, end in message_ranges:
		if end < start:
			continue
		for r in range(start, end + 1):
			for c in range(1, 14 + 1):  # Columns A..N
				left = thin
				right = thin
				top = thin
				bottom = thin

				if c == 1:
					left = thick
				if c == 14:
					right = thick
				if r == start:
					top = thick
				if r == end:
					bottom = thick

				cell = ws.cell(row=r, column=c)
				cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def write_excel(db, template_path: str, out_path: str):
	df_bus = build_main_bus_sheet(db)
	df_msgs = build_main_bus_message_sheet(db)

	if not os.path.exists(template_path):
		print(f"Error: Template Excel file not found: {template_path}")
		sys.exit(1)

	try:
		wb = load_workbook(template_path)
	except Exception as exc:
		print(f"Failed to load Excel template: {exc}")
		sys.exit(1)

	# --- Update 'Main BUS' sheet ---
	if "Main BUS" in wb.sheetnames:
		ws_bus = wb["Main BUS"]
		header_row_bus = _find_header_row(ws_bus, ["ID", "Name"])
		if header_row_bus is None:
			print("Warning: Could not find header row in 'Main BUS' sheet; skipping.")
		else:
			aliases_bus = {
				"Description": ["Discription"],
				"Length (bytes)": ["length (bytes)", "length bytes"],
				"Receiver": ["Reciever", "Reciver"],
			}
			bus_cols = ["ID", "Name", "Description", "Length (bytes)", "Sender", "Receiver"]
			head_map_bus = _map_headers(ws_bus, header_row_bus, bus_cols, aliases_bus)
			rows_bus = df_bus.to_dict(orient="records")
			_populate_sheet(ws_bus, header_row_bus, head_map_bus, rows_bus)
	else:
		print("Warning: 'Main BUS' sheet not found; skipping.")

	# --- Update 'Main BUS Message' sheet ---
	if "Main BUS Message" in wb.sheetnames:
		ws_msg = wb["Main BUS Message"]
		header_row_msg = _find_header_row(ws_msg, ["ID", "Name", "start bit"])
		if header_row_msg is None:
			print("Warning: Could not find header row in 'Main BUS Message' sheet; skipping.")
		else:
			aliases_msg = {
				"Description": ["Discription"],
				"length (bits)": ["length", "length bits"],
			}
			msg_cols = [
				"ID",
				"Name",
				"Description",
				"start bit",
				"length (bits)",
				"Length (bytes)",
				"Gain",
				"Offset",
				"Signed",
				"Byte Order",
				"Unit",
			]
			head_map_msg = _map_headers(ws_msg, header_row_msg, msg_cols, aliases_msg)
			# Convert to list of dicts and, for each message, insert an extra
			# padding row (blank except for start bit) followed by a blank
			# separator row between messages.
			rows_msg_raw = df_msgs.to_dict(orient="records")
			rows_msg_grouped: list[dict[str, object]] = []
			prev_id = None
			last_row_for_id: dict[str, object] | None = None
			for row in rows_msg_raw:
				cur_id = row.get("ID")
				if prev_id is not None and cur_id != prev_id:
					# Close previous message with a padding row and a blank separator
					if last_row_for_id is not None:
						try:
							prev_start = int(last_row_for_id.get("start bit"))
							prev_len = int(last_row_for_id.get("length (bits)"))
							padding_start = prev_start + prev_len
							rows_msg_grouped.append({"__row_mode": "padding", "start bit": padding_start})
						except (TypeError, ValueError):
							pass
						rows_msg_grouped.append({})
				last_row_for_id = row
				rows_msg_grouped.append(row)
				prev_id = cur_id

			# After the loop, add the padding row for the final message
			if last_row_for_id is not None:
				try:
					prev_start = int(last_row_for_id.get("start bit"))
					prev_len = int(last_row_for_id.get("length (bits)"))
					padding_start = prev_start + prev_len
					rows_msg_grouped.append({"__row_mode": "padding", "start bit": padding_start})
				except (TypeError, ValueError):
					pass

			# For each message, only keep the ID and Length (bytes) on the
			# first signal row; padding and separator rows have no ID.
			prev_id = None
			for row in rows_msg_grouped:
				if not row:
					# Separator row resets grouping
					prev_id = None
					continue
				cur_id = row.get("ID")
				if cur_id in (None, ""):
					continue
				if prev_id is None or cur_id != prev_id:
					prev_id = cur_id
				else:
					row["ID"] = ""
					if "Length (bytes)" in row:
						row["Length (bytes)"] = ""

			_populate_sheet(ws_msg, header_row_msg, head_map_msg, rows_msg_grouped)
			_style_message_blocks(ws_msg, header_row_msg, head_map_msg)
	else:
		print("Warning: 'Main BUS Message' sheet not found; skipping.")

	try:
		wb.save(out_path)
	except Exception as exc:
		print(f"Failed to save updated Excel file: {exc}")
		sys.exit(1)


def parse_args() -> str:
	"""Get the DBC path either from CLI or a file picker.

	Behavior:
	- If a path is given on the command line, use it directly.
	- Otherwise, open a file selection dialog to pick a .dbc file.
	- If the dialog is cancelled, exit gracefully.
	"""

	argv = sys.argv[1:]

	if len(argv) >= 1:
		in_dbc = argv[0]
		return os.path.abspath(in_dbc)

	if tk is None or filedialog is None:
		print("Error: Tkinter is not available; please provide the DBC path as a command-line argument.")
		sys.exit(1)

	root = tk.Tk()
	root.withdraw()  # Hide the main window
	root.update()

	file_path = filedialog.askopenfilename(
		title="Select DBC file",
		filetypes=[("DBC files", "*.dbc"), ("All files", "*.*")],
	)
	root.destroy()

	if not file_path:
		print("No DBC file selected. Exiting.")
		sys.exit(0)

	return os.path.abspath(file_path)


def main():
	in_dbc = parse_args()

	# Derive output name: CAN_Loading_<dbcbasename>.xlsx
	dbc_base = os.path.splitext(os.path.basename(in_dbc))[0]
	out_xlsx = os.path.join(SCRIPT_DIR, f"CAN_Loading_{dbc_base}.xlsx")

	print("=== DBC -> XLSX Converter ===")
	print(f"Input DBC : {in_dbc}")
	print(f"Template  : {DEFAULT_TEMPLATE_XLSX}")
	print(f"Output XLSX: {out_xlsx}")

	db = load_dbc(in_dbc)
	write_excel(db, DEFAULT_TEMPLATE_XLSX, out_xlsx)
    
	print("Done.")


if __name__ == "__main__":
	main()

